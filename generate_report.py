"""
LinkedIn Daily Job Report Generator
Mohith Gurram — Data Engineer Job Automation
Scrapes LinkedIn via Apify, scores against resume, generates Excel report, emails it.
"""

import os
import json
import time
import smtplib
import requests
from datetime import datetime, date
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
APIFY_API_TOKEN   = os.environ["APIFY_API_TOKEN"]
EMAIL_SENDER      = os.environ["EMAIL_SENDER"]       # Gmail address used to send
EMAIL_PASSWORD    = os.environ["EMAIL_PASSWORD"]     # Gmail App Password
EMAIL_RECIPIENT   = "mohithgurram03@gmail.com"
ACTOR_ID          = "curious_coder/linkedin-jobs-scraper"
TODAY             = date.today().isoformat()

SEARCH_URLS = [
    "https://www.linkedin.com/jobs/search/?keywords=data%20engineer&location=India&f_TPR=r86400&sortBy=DD",
    "https://www.linkedin.com/jobs/search/?keywords=senior%20data%20engineer&location=India&f_TPR=r86400&sortBy=DD",
    "https://www.linkedin.com/jobs/search/?keywords=azure%20data%20engineer&location=India&f_TPR=r86400&sortBy=DD",
]

# Mohith's resume keywords (weighted)
RESUME_KEYWORDS = [
    "azure", "databricks", "snowflake", "pyspark", "spark", "kafka",
    "airflow", "data factory", "adf", "delta lake", "pipeline", "etl",
    "elt", "python", "sql", "data lake", "adls", "data warehouse",
    "data engineering", "hadoop", "hive", "teradata",
]

EXCLUDE_TITLES = [
    "data analyst", "business intelligence", "fullstack", "machine learning",
    "backend and data", "junior", "software engineer",
]


# ── Step 1: Scrape LinkedIn via Apify ─────────────────────────────────────────
def scrape_linkedin_jobs():
    print("🔍 Scraping LinkedIn jobs via Apify...")
    url = f"https://api.apify.com/v2/acts/{ACTOR_ID}/runs"
    payload = {
        "urls": SEARCH_URLS,
        "count": 100,
        "scrapeCompany": False,
    }
    headers = {"Authorization": f"Bearer {APIFY_API_TOKEN}", "Content-Type": "application/json"}

    # Start run
    resp = requests.post(url, json=payload, headers=headers, timeout=30)
    resp.raise_for_status()
    run_id = resp.json()["data"]["id"]
    print(f"   Run started: {run_id}")

    # Poll until finished
    status_url = f"https://api.apify.com/v2/actor-runs/{run_id}"
    for attempt in range(60):
        time.sleep(10)
        status_resp = requests.get(status_url, headers=headers, timeout=15)
        status = status_resp.json()["data"]["status"]
        print(f"   Status [{attempt+1}]: {status}")
        if status == "SUCCEEDED":
            break
        if status in ("FAILED", "ABORTED", "TIMED-OUT"):
            raise RuntimeError(f"Apify run failed with status: {status}")
    else:
        raise RuntimeError("Apify run timed out after 10 minutes.")

    # Fetch dataset
    dataset_id = status_resp.json()["data"]["defaultDatasetId"]
    dataset_url = f"https://api.apify.com/v2/datasets/{dataset_id}/items?limit=200"
    items_resp = requests.get(dataset_url, headers=headers, timeout=30)
    items_resp.raise_for_status()
    jobs = items_resp.json()
    print(f"   ✅ Fetched {len(jobs)} jobs from Apify.")
    return jobs


# ── Step 2: Score & Filter ────────────────────────────────────────────────────
def score_job(job):
    title = (job.get("title") or "").lower()
    desc  = (job.get("descriptionText") or "").lower()
    combined = title + " " + desc

    for ex in EXCLUDE_TITLES:
        if ex in title:
            return 0

    score = sum(1 for kw in RESUME_KEYWORDS if kw in combined)
    if "senior data engineer" in title or "data engineer" in title or "azure data engineer" in title:
        score += 3
    try:
        applicants = int(job.get("applicantsCount") or 999)
        if applicants < 100:
            score += 2
        elif applicants < 150:
            score += 1
    except (ValueError, TypeError):
        pass
    posted = (job.get("postedAt") or "")
    if posted == TODAY:
        score += 1
    return score


def get_priority(score, applicants):
    try:
        apps = int(applicants or 999)
    except (ValueError, TypeError):
        apps = 999
    if score >= 18 or (score >= 14 and apps < 100):
        return "HIGH"
    elif score >= 12:
        return "MEDIUM"
    return "LOW"


def get_key_skills(job):
    desc = (job.get("descriptionText") or "").lower()
    found = []
    for kw in ["azure", "databricks", "snowflake", "pyspark", "spark", "kafka",
               "airflow", "data factory", "delta lake", "python", "sql",
               "scala", "synapse", "adls", "dbt", "hadoop", "hive", "aws", "gcp"]:
        if kw in desc:
            found.append(kw.upper() if kw in ["sql","adf","adls","aws","gcp","dbt"] else kw.title())
    return ", ".join(found[:8]) if found else "Azure, PySpark, SQL"


def filter_and_rank(jobs):
    seen_links = set()
    scored = []
    for j in jobs:
        link = j.get("link") or j.get("url") or ""
        if link in seen_links:
            continue
        seen_links.add(link)
        s = score_job(j)
        if s > 0:
            scored.append((s, j))
    scored.sort(key=lambda x: -x[0])
    print(f"   ✅ {len(scored)} matching jobs after filtering and ranking.")
    return scored[:30]


# ── Step 3: Generate Excel Report ─────────────────────────────────────────────
def generate_excel(ranked_jobs, output_path):
    print("📊 Generating Excel report...")
    wb = Workbook()

    DARK_BLUE = "1F497D"
    thin = Side(style="thin", color="CCCCCC")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── Sheet 1: Job Listings ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Job Listings"

    # Title
    ws.merge_cells("A1:K1")
    ws["A1"] = f"LinkedIn Daily Job Report — Data Engineer in India | {TODAY}"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:K2")
    ws["A2"] = "Candidate: Mohith Gurram  |  Ranked by Resume Match Score  |  Source: LinkedIn (Last 24 Hours)"
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="555555")
    ws["A2"].fill = PatternFill("solid", fgColor="EBF2FA")
    ws["A2"].alignment = center
    ws.row_dimensions[2].height = 16

    # Headers
    headers    = ["#", "Job Title", "Company", "Location", "Posted", "Applicants", "Priority", "Score", "Key Skills", "Why It Matches", "Apply"]
    col_widths = [4,   30,          24,         26,         12,       12,            10,         8,       38,           48,               14]
    ws.row_dimensions[3].height = 30
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=DARK_BLUE)
        c.alignment = center
        c.border = cell_border
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Data rows
    HIGH_FILL = PatternFill("solid", fgColor="C6EFCE")
    MED_FILL  = PatternFill("solid", fgColor="FFEB9C")
    LOW_FILL  = PatternFill("solid", fgColor="FCE4D6")
    ALT_FILL  = PatternFill("solid", fgColor="F0F5FB")
    WHT_FILL  = PatternFill("solid", fgColor="FFFFFF")

    for i, (score, j) in enumerate(ranked_jobs):
        row = i + 4
        ws.row_dimensions[row].height = 50
        is_alt    = i % 2 == 1
        row_fill  = ALT_FILL if is_alt else WHT_FILL
        applicants = j.get("applicantsCount") or "N/A"
        priority   = get_priority(score, applicants)
        p_fill     = HIGH_FILL if priority == "HIGH" else MED_FILL if priority == "MEDIUM" else LOW_FILL
        p_icon     = "🔥 HIGH" if priority == "HIGH" else "⭐ MED" if priority == "MEDIUM" else "LOW"
        p_color    = "1A7A1A" if priority == "HIGH" else "7D5A00" if priority == "MEDIUM" else "9C0000"
        link       = j.get("link") or ""
        posted     = j.get("postedAt") or ""
        title      = j.get("title") or ""
        company    = j.get("companyName") or ""
        location   = j.get("location") or ""
        skills     = get_key_skills(j)
        desc_short = (j.get("descriptionText") or "")[:220].replace("\n", " ").strip()

        vals = [
            (i+1,      Font(name="Arial", bold=True, size=9, color="555555"), row_fill, center),
            (title,    Font(name="Arial", bold=True, size=9, color=DARK_BLUE), row_fill, left),
            (company,  Font(name="Arial", italic=True, size=9, color="444444"), row_fill, left),
            (location, Font(name="Arial", size=9), row_fill, left),
            (posted,   Font(name="Arial", size=9), row_fill, center),
            (str(applicants), Font(name="Arial", size=9, bold=(str(applicants).isdigit() and int(applicants) < 100)), row_fill, center),
            (p_icon,   Font(name="Arial", bold=True, size=9, color=p_color), p_fill, center),
            (score,    Font(name="Arial", bold=True, size=9), row_fill, center),
            (skills,   Font(name="Arial", size=8, color="1F497D"), row_fill, left),
            (desc_short, Font(name="Arial", size=8, color="333333"), row_fill, left),
        ]
        for ci, (val, font, fill, align) in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = font; c.fill = fill; c.alignment = align; c.border = cell_border

        # Apply link
        lc = ws.cell(row=row, column=11, value="Apply →")
        lc.hyperlink = link
        lc.font = Font(name="Arial", size=9, color="0563C1", underline="single")
        lc.fill = row_fill; lc.alignment = center; lc.border = cell_border

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:K{3 + len(ranked_jobs)}"

    # ── Sheet 2: Dashboard ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Dashboard")
    for col, w in zip(["A","B","C","D","E","F"], [4, 30, 22, 18, 36, 22]):
        ws2.column_dimensions[col].width = w

    ws2.merge_cells("A1:F1")
    ws2["A1"] = f"Job Search Dashboard — Mohith Gurram | {TODAY}"
    ws2["A1"].font = Font(name="Arial", bold=True, size=15, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws2["A1"].alignment = center
    ws2.row_dimensions[1].height = 32

    ws2.merge_cells("A2:F2")
    ws2["A2"] = "Data Engineer / Senior Data Engineer | India | Roles posted in last 24 hours"
    ws2["A2"].font = Font(name="Arial", italic=True, size=10, color="444444")
    ws2["A2"].fill = PatternFill("solid", fgColor="EBF2FA")
    ws2["A2"].alignment = center
    ws2.row_dimensions[2].height = 18

    # Stats
    total      = len(ranked_jobs)
    high_count = sum(1 for s,j in ranked_jobs if get_priority(s, j.get("applicantsCount")) == "HIGH")
    med_count  = sum(1 for s,j in ranked_jobs if get_priority(s, j.get("applicantsCount")) == "MEDIUM")
    low_count  = total - high_count - med_count
    today_cnt  = sum(1 for s,j in ranked_jobs if (j.get("postedAt") or "") == TODAY)

    ws2.row_dimensions[4].height = 20
    ws2.row_dimensions[5].height = 38
    for ci, (lbl, val) in enumerate([
        ("Total Jobs", total), ("🔥 HIGH Priority", high_count),
        ("⭐ MED Priority", med_count), ("LOW Priority", low_count), ("Posted Today", today_cnt)
    ], 2):
        lc = ws2.cell(row=4, column=ci, value=lbl)
        lc.font = Font(name="Arial", size=9, color="888888")
        lc.fill = PatternFill("solid", fgColor="F7F9FC")
        lc.alignment = center; lc.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        vc = ws2.cell(row=5, column=ci, value=val)
        vc.font = Font(name="Arial", bold=True, size=20, color=DARK_BLUE)
        vc.fill = PatternFill("solid", fgColor="F7F9FC")
        vc.alignment = center; vc.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # High priority table
    ws2.row_dimensions[7].height = 22
    ws2.merge_cells("B7:F7")
    ws2["B7"] = "🔥 HIGH Priority Jobs — Apply Immediately"
    ws2["B7"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws2["B7"].fill = PatternFill("solid", fgColor="1A7A1A")
    ws2["B7"].alignment = Alignment(horizontal="left", vertical="center")

    ws2.row_dimensions[8].height = 18
    for ci, h in enumerate(["Company", "Role", "Applicants", "Key Skills", "Apply"], 2):
        c = ws2.cell(row=8, column=ci, value=h)
        c.font = Font(name="Arial", bold=True, size=9, color="333333")
        c.fill = PatternFill("solid", fgColor="C6EFCE")
        c.alignment = center
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    high_jobs = [(s,j) for s,j in ranked_jobs if get_priority(s, j.get("applicantsCount")) == "HIGH"]
    for ri, (s, j) in enumerate(high_jobs, 9):
        ws2.row_dimensions[ri].height = 20
        link = j.get("link") or ""
        row_bg = "F0FBF0" if ri % 2 == 1 else "FFFFFF"
        for ci, val in enumerate([
            j.get("companyName") or "",
            j.get("title") or "",
            str(j.get("applicantsCount") or "N/A"),
            get_key_skills(j),
        ], 2):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Arial", size=9)
            c.fill = PatternFill("solid", fgColor=row_bg)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        lc = ws2.cell(row=ri, column=6, value="Apply →")
        lc.hyperlink = link
        lc.font = Font(name="Arial", size=9, color="0563C1", underline="single")
        lc.fill = PatternFill("solid", fgColor=row_bg)
        lc.alignment = center
        lc.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws2.freeze_panes = "A3"

    wb.save(output_path)
    print(f"   ✅ Excel saved: {output_path}")
    return output_path


# ── Step 4: Send Email ─────────────────────────────────────────────────────────
def send_email(report_path, job_count, high_count):
    print("📧 Sending email...")
    msg = MIMEMultipart()
    msg["From"]    = EMAIL_SENDER
    msg["To"]      = EMAIL_RECIPIENT
    msg["Subject"] = f"🔥 Daily LinkedIn Job Report — {job_count} Data Engineer Jobs | {TODAY}"

    body = f"""
Hi Mohith,

Here is your daily LinkedIn job report for Data Engineer / Senior Data Engineer roles in India posted in the last 24 hours.

📊 Today's Summary ({TODAY}):
  • Total matching jobs found : {job_count}
  • 🔥 HIGH priority jobs      : {high_count}  ← Apply to these first!
  • Score scale                : up to 27 (keyword match + title bonus + applicant count bonus)

📌 Quick Tips:
  • Sort by "Score" in the Excel sheet for strongest matches
  • Filter by "Priority = 🔥 HIGH" for best interview odds
  • Jobs with < 100 applicants are your sweet spot — apply today!

The full report is attached as an Excel file with two sheets:
  1. Job Listings — All {job_count} ranked jobs with clickable Apply links
  2. Dashboard   — Summary stats + HIGH priority shortlist

Good luck today! 🚀

---
This report is auto-generated daily at 10:30 PM IST via GitHub Actions + Apify LinkedIn Scraper.
Resume match keywords: Azure, Databricks, Snowflake, PySpark, Spark, Kafka, Airflow, ADF, Delta Lake, Python, SQL and more.
    """.strip()

    msg.attach(MIMEText(body, "plain"))

    # Attach Excel
    filename = os.path.basename(report_path)
    with open(report_path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename={filename}")
    msg.attach(part)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(EMAIL_SENDER, EMAIL_PASSWORD)
        server.sendmail(EMAIL_SENDER, EMAIL_RECIPIENT, msg.as_string())

    print(f"   ✅ Email sent to {EMAIL_RECIPIENT}")


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    print(f"\n{'='*60}")
    print(f"  LinkedIn Job Report — {TODAY}")
    print(f"{'='*60}\n")

    raw_jobs      = scrape_linkedin_jobs()
    ranked_jobs   = filter_and_rank(raw_jobs)

    if not ranked_jobs:
        print("⚠️  No matching jobs found today. Exiting.")
        return

    report_path = f"/tmp/Mohith_LinkedIn_Jobs_{TODAY}.xlsx"
    generate_excel(ranked_jobs, report_path)

    high_count = sum(1 for s,j in ranked_jobs if get_priority(s, j.get("applicantsCount")) == "HIGH")
    send_email(report_path, len(ranked_jobs), high_count)

    print(f"\n✅ Done! Report with {len(ranked_jobs)} jobs sent to {EMAIL_RECIPIENT}\n")


if __name__ == "__main__":
    main()
